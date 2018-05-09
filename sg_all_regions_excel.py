#Script to create an Excel file with information on all security groups in all regions
#Author : Jino Thomas
#Input : Access key, Secret Key, Session Token
#Note : Make sure that destination file name is correct if you want the output somewhere else other than current working directory

import boto3
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import DARKBLUE
import string

access_key = input("Enter AWS_ACCESS_KEY_ID: ")
secret_key = input("Enter AWS_SECRET_ACCESS_KEY: ")

session = boto3.Session(aws_access_key_id=access_key, aws_secret_access_key=secret_key)

ec2 = session.client('ec2', region_name='us-west-2')
regions = [region['RegionName'] for region in ec2.describe_regions()['Regions']]
wb = Workbook()
dest_filename = './sg_all_details.xlsx'

for region in regions:
    ec2 = session.client('ec2', region_name=region)
    wsht = 'ws_%s' % region
    wsht = wb.create_sheet(title=region)
    font = Font(color=DARKBLUE, bold=True)
    wsht.append(["Description", "Group Name", "Inbound Rules", "Owner Account", "Group ID", "Outbound Rules", "Tags", "VPC Id"])
    for x in string.ascii_uppercase:
        c = wsht['%s1' % x]
        c.font = font
    responseset = ec2.describe_security_groups()
    row = 1
    for item in responseset['SecurityGroups']:
        row += 1
        col = 1
        while col < 8:
            for key, value in item.items():
                _ = wsht.cell(column=col, row=row, value="{0}".format(value))
                col += 1
wb.remove(wb["Sheet"])
wb.save(filename = dest_filename)
